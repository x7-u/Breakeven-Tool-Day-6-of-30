"""Day 6. One-shot generator for the bundled CVP workbooks.

Each workbook ships:
  - metadata sheet (company, currency, period_label, optional current_volume,
    capacity_units, non_cash_amount, industry)
  - unit_economics sheet (fixed_cost, variable_cost_per_unit, price_per_unit,
    optional target_profit)
  - products sheet (optional, for multi-product mode)
  - step_costs sheet (optional, for stepped fixed cost)

Four samples exercise different cost structures:
  1. Pour & Roast Coffee     (GBP) - low fixed, low price, high volume
  2. SkyHopper Airways       (EUR) - huge fixed, thin per-seat margin
  3. Ledgerly SaaS           (USD) - high fixed, fat per-seat margin
  4. Brew & Bites Cafe       (GBP) - multi-product (drinks/food/retail) + step costs

Each rebuild round-trips through parse_inputs() and asserts the headline
break-even matches the expected value to +/- 1 unit before saving.
"""
from __future__ import annotations

import math
import sys
from pathlib import Path

import openpyxl
from openpyxl.styles import Font

HERE = Path(__file__).resolve().parent
DAY_ROOT = HERE.parent
PROJECT_ROOT = DAY_ROOT.parent
for p in (str(DAY_ROOT), str(PROJECT_ROOT)):
    if p not in sys.path:
        sys.path.insert(0, p)

from cvp_schema import parse_inputs


def _kv_sheet(ws, rows: list[tuple[str, object]]) -> None:
    ws.append(["key", "value"])
    ws["A1"].font = Font(bold=True)
    ws["B1"].font = Font(bold=True)
    for k, v in rows:
        ws.append([k, v])


def _table_sheet(ws, headers: list[str], rows: list[list[object]]) -> None:
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True)
    for r in rows:
        ws.append(r)


def _build(out_path: Path, *,
           company: str, currency: str, period_label: str,
           fixed_cost: float, variable_cost_per_unit: float,
           price_per_unit: float,
           current_volume: float | None = None,
           target_profit: float | None = None,
           capacity_units: float | None = None,
           non_cash_amount: float | None = None,
           industry: str | None = None,
           covenant_min_revenue: float | None = None,
           products: list[tuple[str, float, float, float]] | None = None,
           step_costs: list[tuple[float, float]] | None = None,
           volume_curve: list[tuple[str, float]] | None = None) -> None:
    wb = openpyxl.Workbook()
    md = wb.active
    md.title = "metadata"
    md_rows: list[tuple[str, object]] = [
        ("company", company),
        ("currency", currency),
        ("period_label", period_label),
    ]
    if current_volume is not None:
        md_rows.append(("current_volume", current_volume))
    if capacity_units is not None:
        md_rows.append(("capacity_units", capacity_units))
    if non_cash_amount is not None:
        md_rows.append(("non_cash_amount", non_cash_amount))
    if industry is not None:
        md_rows.append(("industry", industry))
    if covenant_min_revenue is not None:
        md_rows.append(("covenant_min_revenue", covenant_min_revenue))
    _kv_sheet(md, md_rows)

    ue = wb.create_sheet("unit_economics")
    ue_rows: list[tuple[str, object]] = [
        ("fixed_cost", fixed_cost),
        ("variable_cost_per_unit", variable_cost_per_unit),
        ("price_per_unit", price_per_unit),
    ]
    if target_profit is not None:
        ue_rows.append(("target_profit", target_profit))
    _kv_sheet(ue, ue_rows)

    if products:
        prod = wb.create_sheet("products")
        _table_sheet(
            prod,
            ["name", "price_per_unit", "variable_cost_per_unit", "mix_pct"],
            [list(p) for p in products],
        )
    if step_costs:
        sc = wb.create_sheet("step_costs")
        _table_sheet(
            sc,
            ["above_units", "extra_fixed_cost"],
            [list(s) for s in step_costs],
        )
    if volume_curve:
        vc = wb.create_sheet("volume_curve")
        _table_sheet(
            vc,
            ["period_label", "units"],
            [list(v) for v in volume_curve],
        )

    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)

    parsed = parse_inputs(path=out_path, source_filename=out_path.name)
    cm = parsed.inputs.price_per_unit - parsed.inputs.variable_cost_per_unit
    expected_be = fixed_cost / (price_per_unit - variable_cost_per_unit)
    actual_be = parsed.inputs.fixed_cost / cm
    assert math.isclose(expected_be, actual_be, abs_tol=1.0), (
        f"BE mismatch in {out_path.name}: expected {expected_be}, got {actual_be}"
    )


def build_coffee_shop(path: Path) -> None:
    _build(
        path,
        company="Pour & Roast Coffee",
        currency="GBP",
        period_label="2026-Q2",
        fixed_cost=18_000.0,
        variable_cost_per_unit=1.10,
        price_per_unit=3.20,
        current_volume=11_000,
        target_profit=4_000.0,
        capacity_units=15_000,           # max cups/month at current staff
        non_cash_amount=2_000.0,         # equipment depreciation
        industry="coffee shop",
    )


def build_low_cost_airline(path: Path) -> None:
    _build(
        path,
        company="SkyHopper Airways",
        currency="EUR",
        period_label="2026-Q2",
        fixed_cost=8_500_000.0,
        variable_cost_per_unit=67.00,
        price_per_unit=89.00,
        current_volume=420_000,
        target_profit=1_500_000.0,
        capacity_units=480_000,          # fleet + slots ceiling
        non_cash_amount=1_400_000.0,     # fleet depreciation
        industry="airline",
    )


def build_software_saas(path: Path) -> None:
    _build(
        path,
        company="Ledgerly SaaS",
        currency="USD",
        period_label="2026-04",
        fixed_cost=240_000.0,
        variable_cost_per_unit=4.00,
        price_per_unit=49.00,
        current_volume=4_800,
        target_profit=60_000.0,
        capacity_units=15_000,           # support team can handle this many seats
        non_cash_amount=18_000.0,        # capitalised software amortisation
        industry="saas",
    )


def build_brew_and_bites(path: Path) -> None:
    """Multi-product cafe with a 6-month ramp-up plan and a bank covenant.
    Headline UE uses the blended average; the products sheet overrides for
    the multi-product BE. Step costs reflect a second barista hire."""
    _build(
        path,
        company="Brew & Bites Cafe",
        currency="GBP",
        period_label="2026-Q2",
        fixed_cost=24_000.0,
        variable_cost_per_unit=1.55,     # blended VC across the mix
        price_per_unit=4.20,             # blended price across the mix
        current_volume=10_500,           # transactions/month
        target_profit=6_000.0,
        capacity_units=18_000,
        non_cash_amount=3_000.0,
        industry="cafe",
        covenant_min_revenue=42_000.0,   # bank requires GBP 42k minimum monthly revenue
        products=[
            ("Espresso drinks",          3.40,  1.00, 0.55),
            ("Pastries & food",          4.80,  2.20, 0.30),
            ("Retail bag & merch",       9.50,  4.50, 0.15),
        ],
        step_costs=[
            (12_000, 4_500.0),
            (16_000, 3_500.0),
        ],
        volume_curve=[
            # 6-month ramp; opens at 60% of run-rate and grows
            ("2026-04",  6_300.0),
            ("2026-05",  8_200.0),
            ("2026-06",  9_400.0),
            ("2026-07", 10_500.0),
            ("2026-08", 11_300.0),
            ("2026-09", 12_400.0),
        ],
    )


def main() -> None:
    out_dir = HERE
    build_coffee_shop(out_dir / "sample_coffee_shop.xlsx")
    build_low_cost_airline(out_dir / "sample_low_cost_airline.xlsx")
    build_software_saas(out_dir / "sample_software_saas.xlsx")
    build_brew_and_bites(out_dir / "sample_brew_and_bites.xlsx")
    print("Wrote 4 Day 6 sample workbooks.")


if __name__ == "__main__":
    main()
