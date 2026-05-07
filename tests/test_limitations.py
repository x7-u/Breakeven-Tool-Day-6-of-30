"""Day 6. Tests for the 4 limitation-fix features:
  - Monte Carlo simulation
  - Time-phased BE
  - Bank covenant overlay
  - volume_curve sheet parsing
"""
from __future__ import annotations

import io
import math
from pathlib import Path

import openpyxl
import pytest
from cvp_maths import (
    first_crossing_period,
    headline_stats,
    time_phased_be,
)
from cvp_schema import CVPInputs, CVPMetadata, VolumePoint, parse_inputs
from monte_carlo import simulate as mc_simulate

# ---- Monte Carlo --------------------------------------------------

def test_monte_carlo_runs_and_returns_sane_percentiles():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    res = mc_simulate(inp, n_runs=2000, span_pct=0.20, seed=42)
    assert res.n_runs == 2000
    # Percentiles should be ordered
    assert res.p5_be <= res.p25_be <= res.median_be <= res.p75_be <= res.p95_be
    # Median should be roughly the deterministic BE (1000/3 = 333.33)
    assert math.isclose(res.median_be, 333.0, abs_tol=80)
    # No undefined trials for this clean spread
    assert res.pct_undefined == 0.0
    assert len(res.histogram) > 0


def test_monte_carlo_seed_is_deterministic():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    a = mc_simulate(inp, n_runs=1000, span_pct=0.10, seed=7)
    b = mc_simulate(inp, n_runs=1000, span_pct=0.10, seed=7)
    assert a.median_be == b.median_be
    assert a.p5_be == b.p5_be


def test_monte_carlo_handles_high_undefined_when_cm_thin():
    # CM is razor thin (price 5, vc 4.95); +/-20% will frequently flip CM negative
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=4.95, price_per_unit=5.0)
    res = mc_simulate(inp, n_runs=2000, span_pct=0.20, seed=42)
    assert res.pct_undefined > 0.10
    # When some trials produce undefined BE, the rest still summarise
    assert res.median_be > 0


def test_monte_carlo_clamps_extreme_params():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    res = mc_simulate(inp, n_runs=1, span_pct=0.0, seed=1)
    # n_runs gets bumped to 100, span gets bumped to 0.05
    assert res.n_runs >= 100
    assert res.span_pct >= 0.05


# ---- Time-phased BE -----------------------------------------------

def test_time_phased_be_crosses_when_cumulative_profit_goes_positive():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    # cm = 3; need cumulative units > 1000/3 each period (fixed accumulates too)
    # Period 1: 100 units -> cum_profit = 300 - 1000 = -700
    # Period 2: 500 units -> cum_units 600, cum_profit = 1800 - 2000 = -200
    # Period 3: 800 units -> cum_units 1400, cum_profit = 4200 - 3000 = +1200 (crossed)
    curve = [
        VolumePoint("M1", 100),
        VolumePoint("M2", 500),
        VolumePoint("M3", 800),
    ]
    rows = time_phased_be(inp, curve)
    assert len(rows) == 3
    assert rows[0].crossed is False
    assert rows[1].crossed is False
    assert rows[2].crossed is True
    assert first_crossing_period(rows) == "M3"


def test_time_phased_be_never_crosses_when_volumes_too_low():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    curve = [VolumePoint(f"M{i}", 50) for i in range(1, 5)]
    rows = time_phased_be(inp, curve)
    assert all(r.crossed is False for r in rows)
    assert first_crossing_period(rows) is None


def test_time_phased_be_empty_returns_empty():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    assert time_phased_be(inp, []) == []


def test_time_phased_be_handles_negative_cm():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=10, price_per_unit=8)
    curve = [VolumePoint("M1", 100), VolumePoint("M2", 100)]
    rows = time_phased_be(inp, curve)
    # Profit is -fixed * period_index (cm <= 0 contributes nothing)
    assert rows[0].cumulative_profit == -1000
    assert rows[1].cumulative_profit == -2000
    assert all(not r.crossed for r in rows)


# ---- Covenant -----------------------------------------------------

def test_covenant_no_breach_when_revenue_above_floor():
    md = CVPMetadata("X", "GBP", "P", current_volume=1000,
                     covenant_min_revenue=4000)
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    h = headline_stats(inp, md)
    # Current revenue = 1000 * 5 = 5000 > 4000 -> no breach
    assert h.covenant_min_revenue == 4000
    assert h.covenant_breach is False
    assert h.covenant_buffer_pct is not None and h.covenant_buffer_pct > 0


def test_covenant_breach_when_revenue_below_floor():
    md = CVPMetadata("X", "GBP", "P", current_volume=500,
                     covenant_min_revenue=4000)
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    h = headline_stats(inp, md)
    # Current revenue = 500 * 5 = 2500 < 4000 -> breach
    assert h.covenant_breach is True
    assert h.covenant_buffer_pct < 0


def test_covenant_falls_back_to_be_revenue_without_current_volume():
    md = CVPMetadata("X", "GBP", "P", covenant_min_revenue=2000)
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    h = headline_stats(inp, md)
    # BE rev = (1000/3) * 5 = ~1666 < 2000 -> breach
    assert h.covenant_breach is True


def test_covenant_absent_when_not_set():
    md = CVPMetadata("X", "GBP", "P", current_volume=1000)
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    h = headline_stats(inp, md)
    assert h.covenant_min_revenue is None
    assert h.covenant_breach is None


# ---- Schema parsing for volume_curve and covenant ----------------

def _book_with_curve(volume_curve_rows, covenant=None) -> bytes:
    wb = openpyxl.Workbook()
    md = wb.active
    md.title = "metadata"
    md.append(["key", "value"])
    md.append(["company", "X"])
    md.append(["currency", "GBP"])
    md.append(["period_label", "P"])
    md.append(["current_volume", 1000])
    if covenant is not None:
        md.append(["covenant_min_revenue", covenant])
    ue = wb.create_sheet("unit_economics")
    ue.append(["key", "value"])
    ue.append(["fixed_cost", 1000])
    ue.append(["variable_cost_per_unit", 2])
    ue.append(["price_per_unit", 5])
    if volume_curve_rows is not None:
        vc = wb.create_sheet("volume_curve")
        vc.append(["period_label", "units"])
        for row in volume_curve_rows:
            vc.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_schema_reads_volume_curve_sheet():
    data = _book_with_curve([("2026-04", 200), ("2026-05", 400), ("2026-06", 600)])
    parsed = parse_inputs(file_bytes=data)
    assert len(parsed.volume_curve) == 3
    assert parsed.volume_curve[0].period_label == "2026-04"
    assert parsed.volume_curve[2].units == 600


def test_schema_reads_covenant_min_revenue():
    data = _book_with_curve(None, covenant=5000)
    parsed = parse_inputs(file_bytes=data)
    assert parsed.metadata.covenant_min_revenue == 5000


def test_schema_volume_curve_skips_bad_rows():
    data = _book_with_curve([
        ("2026-04", 200),
        ("", 100),               # missing label
        ("2026-05", "not-num"),  # bad units
        ("2026-06", -50),        # negative units
        ("2026-07", 400),
    ])
    parsed = parse_inputs(file_bytes=data)
    # Should keep only the 2 valid rows
    labels = [v.period_label for v in parsed.volume_curve]
    assert labels == ["2026-04", "2026-07"]


def test_schema_volume_curve_warns_on_bad_units():
    data = _book_with_curve([
        ("2026-04", 200),
        ("2026-05", "abc"),
    ])
    parsed = parse_inputs(file_bytes=data)
    assert any("volume_curve" in w and "not numeric" in w for w in parsed.warnings)


# ---- Brew sample exercises all four limitation-fix features ------

def test_brew_sample_has_monte_carlo_and_time_phased_and_covenant():
    from pipeline import analyse, to_dict

    sample = (Path(__file__).resolve().parent.parent
              / "sample_data" / "sample_brew_and_bites.xlsx")
    if not sample.exists():
        pytest.skip("brew_and_bites sample missing")
    res = analyse(path=sample, source_filename=sample.name, skip_ai=True)
    d = to_dict(res)
    # Monte Carlo
    assert d["monte_carlo"]["n_runs"] == 4000
    assert d["monte_carlo"]["median_be"] > 0
    assert len(d["monte_carlo"]["histogram"]) == 24
    # Time-phased
    assert len(d["time_phased"]) == 6
    assert d["first_crossing_period"] is not None
    # Covenant
    assert d["headline"]["covenant_min_revenue"] == 42_000
    assert d["headline"]["covenant_breach"] is not None
