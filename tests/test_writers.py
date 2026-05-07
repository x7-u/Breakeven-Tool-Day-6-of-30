"""Day 6. Tests for cvp_excel + cvp_csv writers."""
from __future__ import annotations

from pathlib import Path

import openpyxl
import pytest
from cvp_csv import COLUMNS, write_csv
from cvp_excel import write_workbook
from cvp_maths import CVPResult, headline_stats, sensitivity_table, tornado
from cvp_schema import CVPInputs, CVPMetadata


def _make_result(currency: str = "GBP", with_target: bool = True,
                 with_volume: bool = True) -> CVPResult:
    md = CVPMetadata("Acme Ltd", currency, "2026-Q2",
                     current_volume=500 if with_volume else None)
    inp = CVPInputs(
        fixed_cost=1000,
        variable_cost_per_unit=2.0,
        price_per_unit=5.0,
        target_profit=600 if with_target else None,
    )
    h = headline_stats(inp, md)
    sens = sensitivity_table(inp)
    tor = tornado(sens)
    return CVPResult(metadata=md, inputs=inp, headline=h, sensitivity=sens, tornado=tor)


def test_excel_three_sheets(tmp_path: Path):
    out = tmp_path / "out.xlsx"
    write_workbook(_make_result(), out)
    wb = openpyxl.load_workbook(out)
    assert set(wb.sheetnames) == {"Summary", "Inputs", "Sensitivity"}


def test_excel_currency_format_applied(tmp_path: Path):
    out = tmp_path / "usd.xlsx"
    write_workbook(_make_result(currency="USD"), out)
    wb = openpyxl.load_workbook(out)
    summary = wb["Summary"]
    # Walk a few currency-format cells; one of them must use the $ format.
    formats = [c.number_format for row in summary.iter_rows(min_row=3, max_row=10,
                                                            min_col=2, max_col=2)
               for c in row]
    assert any('"$"' in fmt for fmt in formats)


def test_excel_handles_missing_optional_fields(tmp_path: Path):
    out = tmp_path / "no_opts.xlsx"
    write_workbook(_make_result(with_target=False, with_volume=False), out)
    wb = openpyxl.load_workbook(out)
    assert "Summary" in wb.sheetnames


def test_csv_round_trip_columns(tmp_path: Path):
    out = tmp_path / "sens.csv"
    write_csv(_make_result(), out)
    text = out.read_text(encoding="utf-8-sig")
    lines = text.splitlines()
    header = lines[0].split(",")
    assert tuple(header) == COLUMNS
    # 3 vars * 5 deltas = 15 data rows
    assert len(lines) - 1 == 15


def test_csv_blank_for_undefined_be(tmp_path: Path):
    """When CM goes to zero under shock, BE columns should be empty (not 'None')."""
    md = CVPMetadata("X", "GBP", "P")
    # price = 2.5 means -20% takes price to vc; BE undefined.
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2.0, price_per_unit=2.5)
    h = headline_stats(inp, md)
    sens = sensitivity_table(inp)
    tor = tornado(sens)
    res = CVPResult(metadata=md, inputs=inp, headline=h, sensitivity=sens, tornado=tor)
    out = tmp_path / "x.csv"
    write_csv(res, out)
    text = out.read_text(encoding="utf-8-sig")
    # No literal "None" should appear in the file.
    assert "None" not in text
