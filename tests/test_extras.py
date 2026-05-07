"""Day 6. Tests for the new feature surface (multi-product, step-cost,
cash BE, capacity, benchmarks, heatmap, scenario)."""
from __future__ import annotations

import io
import math

import openpyxl
import pytest
from benchmarks import all_industries, lookup
from cvp_maths import (
    first_reachable_step_be,
    headline_stats,
    heatmap_grid,
    multi_product_stats,
    step_cost_be,
)
from cvp_schema import CVPInputs, CVPMetadata, Product, StepCost, parse_inputs
from pipeline import scenario

# ---- Multi-product ------------------------------------------------

def test_multi_product_weighted_cm():
    products = [
        Product("A", price_per_unit=4.0, variable_cost_per_unit=1.5, mix_pct=0.6),
        Product("B", price_per_unit=10.0, variable_cost_per_unit=4.0, mix_pct=0.4),
    ]
    mp = multi_product_stats(products, fixed_cost=5000)
    # weighted price = 0.6*4 + 0.4*10 = 6.4
    # weighted vc    = 0.6*1.5 + 0.4*4 = 2.5
    # weighted cm    = 3.9
    assert math.isclose(mp.weighted_price, 6.4, abs_tol=1e-6)
    assert math.isclose(mp.weighted_variable_cost, 2.5, abs_tol=1e-6)
    assert math.isclose(mp.weighted_cm_per_unit, 3.9, abs_tol=1e-6)
    assert math.isclose(mp.blended_break_even_units, 5000 / 3.9, rel_tol=1e-6)
    assert len(mp.per_product) == 2


def test_multi_product_returns_none_for_empty():
    assert multi_product_stats([], fixed_cost=1000) is None


# ---- Step-cost ladder --------------------------------------------

def test_step_cost_ladder_segments():
    inp = CVPInputs(fixed_cost=10_000, variable_cost_per_unit=2, price_per_unit=5)
    steps = [
        StepCost(above_units=2_000, extra_fixed_cost=2_000),
        StepCost(above_units=5_000, extra_fixed_cost=3_000),
    ]
    rows = step_cost_be(inp, steps)
    assert len(rows) == 3
    assert rows[0].fixed_cost_at_segment == 10_000
    assert rows[1].fixed_cost_at_segment == 12_000
    assert rows[2].fixed_cost_at_segment == 15_000
    # baseline BE = 10000/3 = ~3333; falls into segment 1 (2000-5000)
    assert rows[0].is_reachable_in_segment is False
    assert rows[1].is_reachable_in_segment is True
    assert math.isclose(rows[1].break_even_units, 12_000 / 3, abs_tol=1.0)


def test_step_cost_no_steps_returns_one_segment():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    rows = step_cost_be(inp, [])
    assert len(rows) == 1
    assert rows[0].is_reachable_in_segment is True
    assert math.isclose(rows[0].break_even_units, 1000 / 3, abs_tol=1.0)


def test_first_reachable_step_be_helper():
    inp = CVPInputs(fixed_cost=10_000, variable_cost_per_unit=2, price_per_unit=5)
    steps = [
        StepCost(above_units=2_000, extra_fixed_cost=2_000),
        StepCost(above_units=5_000, extra_fixed_cost=3_000),
    ]
    rows = step_cost_be(inp, steps)
    be = first_reachable_step_be(rows)
    assert be is not None
    assert math.isclose(be, 12_000 / 3, abs_tol=1.0)


# ---- Cash BE + capacity ------------------------------------------

def test_cash_break_even_excludes_non_cash():
    md = CVPMetadata("X", "GBP", "P", non_cash_amount=2000)
    inp = CVPInputs(fixed_cost=10_000, variable_cost_per_unit=2, price_per_unit=5)
    h = headline_stats(inp, md)
    assert math.isclose(h.cash_break_even_units, 8000 / 3, rel_tol=1e-6)


def test_capacity_reachable_flag():
    md = CVPMetadata("X", "GBP", "P", capacity_units=5_000)
    inp = CVPInputs(fixed_cost=9_000, variable_cost_per_unit=2, price_per_unit=5)
    h = headline_stats(inp, md)
    # BE = 9000/3 = 3000, capacity 5000 -> reachable
    assert h.capacity_reachable is True
    assert h.capacity_buffer_pct is not None
    assert h.capacity_buffer_pct > 0


def test_capacity_not_reachable_when_be_exceeds():
    md = CVPMetadata("X", "GBP", "P", capacity_units=2_000)
    inp = CVPInputs(fixed_cost=9_000, variable_cost_per_unit=2, price_per_unit=5)
    h = headline_stats(inp, md)
    assert h.capacity_reachable is False


# ---- Benchmarks --------------------------------------------------

def test_benchmark_lookup_known_industry():
    b = lookup("saas")
    assert b is not None
    assert b.cm_ratio_low > 0 and b.cm_ratio_high > b.cm_ratio_low


def test_benchmark_lookup_loose_match():
    b = lookup("Coffee Shop & Cafe")
    assert b is not None
    assert "Coffee" in b.industry


def test_benchmark_lookup_unknown():
    assert lookup(None) is None
    assert lookup("nonsense industry") is None


def test_all_industries_returns_unique_sorted():
    names = all_industries()
    assert names == sorted(names)
    assert len(names) == len(set(names))


# ---- Heatmap -----------------------------------------------------

def test_heatmap_centre_cell_matches_baseline():
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    grid = heatmap_grid(inp, n=5, span=0.20)
    mid = len(grid["deltas"]) // 2
    # centre (0% on price, 0% on vc) should equal baseline BE
    centre = grid["be_grid"][mid][mid]
    assert math.isclose(centre, 1000 / 3, rel_tol=1e-6)


def test_heatmap_handles_negative_cm():
    # Tiny price + 20% drop on price; vc gets bigger -> cm goes negative
    inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=4.5, price_per_unit=5)
    grid = heatmap_grid(inp, n=3, span=0.20)
    flat = [c for row in grid["be_grid"] for c in row]
    assert None in flat  # at least one cell has CM <= 0


# ---- Scenario (live what-if) -------------------------------------

def test_scenario_recomputes_with_overrides():
    base_inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    base_md = CVPMetadata("X", "GBP", "P", current_volume=500)
    out = scenario(base_inputs=base_inp, base_metadata=base_md,
                   overrides={"price_per_unit": 6.0})
    assert out["headline"]["break_even_units"] == 1000 / 4  # cm = 6 - 2 = 4


def test_scenario_keeps_unchanged_fields():
    base_inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    base_md = CVPMetadata("X", "GBP", "P")
    out = scenario(base_inputs=base_inp, base_metadata=base_md, overrides={})
    assert out["inputs"]["price_per_unit"] == 5
    assert out["inputs"]["fixed_cost"] == 1000


def test_scenario_rejects_invalid_inputs():
    base_inp = CVPInputs(fixed_cost=1000, variable_cost_per_unit=2, price_per_unit=5)
    base_md = CVPMetadata("X", "GBP", "P")
    out = scenario(base_inputs=base_inp, base_metadata=base_md,
                   overrides={"price_per_unit": 0})
    assert "error" in out


# ---- Schema parsing for new sheets -------------------------------

def _book(metadata_rows, ue_rows, products=None, step_costs=None) -> bytes:
    wb = openpyxl.Workbook()
    md = wb.active
    md.title = "metadata"
    md.append(["key", "value"])
    for k, v in metadata_rows:
        md.append([k, v])
    ue = wb.create_sheet("unit_economics")
    ue.append(["key", "value"])
    for k, v in ue_rows:
        ue.append([k, v])
    if products:
        prod = wb.create_sheet("products")
        prod.append(["name", "price_per_unit", "variable_cost_per_unit", "mix_pct"])
        for row in products:
            prod.append(list(row))
    if step_costs:
        sc = wb.create_sheet("step_costs")
        sc.append(["above_units", "extra_fixed_cost"])
        for row in step_costs:
            sc.append(list(row))
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_schema_reads_capacity_and_non_cash_and_industry():
    data = _book(
        [("company", "X"), ("currency", "GBP"), ("period_label", "P"),
         ("capacity_units", 5000), ("non_cash_amount", 500), ("industry", "saas")],
        [("fixed_cost", 1000), ("variable_cost_per_unit", 2), ("price_per_unit", 5)],
    )
    parsed = parse_inputs(file_bytes=data)
    assert parsed.metadata.capacity_units == 5000
    assert parsed.metadata.non_cash_amount == 500
    assert parsed.metadata.industry == "saas"


def test_schema_reads_products_sheet_and_normalises_mix():
    data = _book(
        [("company", "X"), ("currency", "GBP"), ("period_label", "P")],
        [("fixed_cost", 1000), ("variable_cost_per_unit", 2), ("price_per_unit", 5)],
        products=[
            ("A", 4.0, 1.5, 0.5),
            ("B", 10.0, 4.0, 0.5),
        ],
    )
    parsed = parse_inputs(file_bytes=data)
    assert len(parsed.products) == 2
    assert math.isclose(sum(p.mix_pct for p in parsed.products), 1.0, abs_tol=1e-6)


def test_schema_normalises_mix_when_not_summing_to_one():
    data = _book(
        [("company", "X"), ("currency", "GBP"), ("period_label", "P")],
        [("fixed_cost", 1000), ("variable_cost_per_unit", 2), ("price_per_unit", 5)],
        products=[
            ("A", 4.0, 1.5, 60),   # not in [0,1]; sums to 100 with B
            ("B", 10.0, 4.0, 40),
        ],
    )
    parsed = parse_inputs(file_bytes=data)
    assert math.isclose(sum(p.mix_pct for p in parsed.products), 1.0, abs_tol=1e-6)
    assert any("auto-normalised" in w for w in parsed.warnings)


def test_schema_reads_step_costs_sheet():
    data = _book(
        [("company", "X"), ("currency", "GBP"), ("period_label", "P")],
        [("fixed_cost", 1000), ("variable_cost_per_unit", 2), ("price_per_unit", 5)],
        step_costs=[(2000, 500), (5000, 1000)],
    )
    parsed = parse_inputs(file_bytes=data)
    assert len(parsed.step_costs) == 2
    assert parsed.step_costs[0].above_units == 2000
    assert parsed.step_costs[1].extra_fixed_cost == 1000


def test_schema_ignores_products_with_missing_columns():
    """If the products sheet is malformed, log a warning and proceed."""
    wb = openpyxl.Workbook()
    md = wb.active
    md.title = "metadata"
    md.append(["key", "value"])
    md.append(["company", "X"])
    md.append(["currency", "GBP"])
    md.append(["period_label", "P"])
    ue = wb.create_sheet("unit_economics")
    ue.append(["key", "value"])
    ue.append(["fixed_cost", 1000])
    ue.append(["variable_cost_per_unit", 2])
    ue.append(["price_per_unit", 5])
    prod = wb.create_sheet("products")
    prod.append(["name", "price"])  # missing the other required cols
    prod.append(["A", 4])
    buf = io.BytesIO()
    wb.save(buf)
    parsed = parse_inputs(file_bytes=buf.getvalue())
    assert parsed.products == []
    assert any("products sheet missing" in w for w in parsed.warnings)


# ---- Round trip on the brew_and_bites sample --------------------

def test_brew_sample_has_products_and_step_costs(tmp_path):
    """Exercises the full pipeline against the sample; should produce both
    multi-product and step-cost outputs."""
    from pathlib import Path

    from pipeline import analyse, to_dict

    sample = (Path(__file__).resolve().parent.parent
              / "sample_data" / "sample_brew_and_bites.xlsx")
    if not sample.exists():
        pytest.skip("multi-product sample missing")
    res = analyse(path=sample, source_filename=sample.name, skip_ai=True)
    d = to_dict(res)
    assert d["multi_product"] is not None
    assert len(d["multi_product"]["per_product"]) == 3
    assert len(d["step_cost_be"]) >= 2
    assert d["benchmark"] is not None
    assert d["headline"]["capacity_reachable"] is not None
    assert d["headline"]["cash_break_even_units"] is not None
    assert d["heatmap"] is not None
