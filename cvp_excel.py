"""Day 6. Excel writer for the BREAK CVP tool.

Three sheets:
  1. Summary: KPIs (BE units / revenue, CM, MoS, target profit, leverage)
     plus the embedded break-even chart PNG.
  2. Inputs: parsed metadata + unit economics echoed back for audit.
  3. Sensitivity: 15-row table of single-variable shocks (3 vars x 5 deltas)
     plus the embedded tornado chart PNG.

Currency-aware formatting (GBP / USD / EUR / fallback).
"""
from __future__ import annotations

from io import BytesIO
from pathlib import Path

from cvp_chart import render_breakeven_png, render_tornado_png
from cvp_maths import SENSITIVITY_VARIABLES, CVPResult
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

_CURRENCY_FMT = {
    "GBP": '"£"#,##0.00;[Red]-"£"#,##0.00',
    "USD": '"$"#,##0.00;[Red]-"$"#,##0.00',
    "EUR": '"€"#,##0.00;[Red]-"€"#,##0.00',
}
INT_FMT = "#,##0"
PCT_FMT = "0.00%"

INK_FILL    = PatternFill("solid", fgColor="0F1117")
PURPLE_FILL = PatternFill("solid", fgColor="7C3AED")
PANEL_FILL  = PatternFill("solid", fgColor="F7F8FA")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
LABEL_FONT  = Font(bold=True, color="0F1117", size=10)
TITLE_FONT  = Font(bold=True, color="0F1117", size=14)


def _ccy_fmt(currency: str) -> str:
    return _CURRENCY_FMT.get(currency.upper(), '#,##0.00')


def write_workbook(result: CVPResult, out_path: Path) -> Path:
    wb = Workbook()
    ws_summary = wb.active
    ws_summary.title = "Summary"
    _write_summary(ws_summary, result)
    _write_inputs(wb.create_sheet("Inputs"), result)
    _write_sensitivity(wb.create_sheet("Sensitivity"), result)
    out_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(out_path)
    return out_path


def _write_summary(ws, result: CVPResult) -> None:
    md = result.metadata
    inp = result.inputs
    h = result.headline
    fmt = _ccy_fmt(md.currency)

    ws["A1"] = f"BREAK | {md.company} | {md.period_label}"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:D1")
    ws["A1"].alignment = Alignment(vertical="center")

    rows: list[tuple[str, object, str]] = [
        ("Currency", md.currency, ""),
        ("Price per unit", inp.price_per_unit, fmt),
        ("Variable cost per unit", inp.variable_cost_per_unit, fmt),
        ("Fixed cost", inp.fixed_cost, fmt),
        ("Contribution margin per unit", h.contribution_margin_per_unit, fmt),
        ("CM ratio", h.cm_ratio, PCT_FMT),
        ("Break-even units", h.break_even_units, INT_FMT),
        ("Break-even revenue", h.break_even_revenue, fmt),
    ]
    if h.target_profit_units is not None:
        rows.append(("Target profit units", h.target_profit_units, INT_FMT))
        rows.append(("Target profit revenue", h.target_profit_revenue, fmt))
    if md.current_volume is not None:
        rows.append(("Current volume", md.current_volume, INT_FMT))
    if h.margin_of_safety_units is not None:
        rows.append(("Margin of safety (units)", h.margin_of_safety_units, INT_FMT))
    if h.margin_of_safety_pct is not None:
        rows.append(("Margin of safety (%)", h.margin_of_safety_pct, PCT_FMT))
    if h.operating_leverage_at_current is not None:
        rows.append(("Operating leverage (current)", h.operating_leverage_at_current, "0.00"))

    start = 3
    for i, (label, value, num_fmt) in enumerate(rows):
        r = start + i
        ws.cell(row=r, column=1, value=label).font = LABEL_FONT
        c = ws.cell(row=r, column=2, value=value)
        if num_fmt:
            c.number_format = num_fmt
        if i % 2 == 0:
            ws.cell(row=r, column=1).fill = PANEL_FILL
            ws.cell(row=r, column=2).fill = PANEL_FILL

    ws.column_dimensions["A"].width = 32
    ws.column_dimensions["B"].width = 22

    # Embed break-even chart.
    png_bytes = render_breakeven_png(
        inp,
        title=f"Break-even chart {md.company}",
        currency=md.currency,
        current_volume=md.current_volume,
    )
    img = XLImage(BytesIO(png_bytes))
    img.width, img.height = 720, 330
    ws.add_image(img, "D3")


def _write_inputs(ws, result: CVPResult) -> None:
    md = result.metadata
    inp = result.inputs
    pairs: list[tuple[str, object]] = [
        ("Company", md.company),
        ("Currency", md.currency),
        ("Period", md.period_label),
        ("Current volume", md.current_volume if md.current_volume is not None else ""),
        ("Fixed cost", inp.fixed_cost),
        ("Variable cost per unit", inp.variable_cost_per_unit),
        ("Price per unit", inp.price_per_unit),
        ("Target profit", inp.target_profit if inp.target_profit is not None else ""),
    ]
    ws["A1"] = "Inputs (echoed for audit)"
    ws["A1"].font = TITLE_FONT
    ws.merge_cells("A1:B1")

    fmt = _ccy_fmt(md.currency)
    for i, (k, v) in enumerate(pairs, start=3):
        ws.cell(row=i, column=1, value=k).font = LABEL_FONT
        c = ws.cell(row=i, column=2, value=v)
        if isinstance(v, (int, float)) and k not in {"Current volume"}:
            c.number_format = fmt
        elif k == "Current volume" and isinstance(v, (int, float)):
            c.number_format = INT_FMT
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 22


def _write_sensitivity(ws, result: CVPResult) -> None:
    md = result.metadata
    fmt = _ccy_fmt(md.currency)

    headers = ["Variable", "Delta %", "New value", "New BE units", "Swing units", "Swing %"]
    for col, name in enumerate(headers, start=1):
        c = ws.cell(row=1, column=col, value=name)
        c.fill = INK_FILL
        c.font = HEADER_FONT
        c.alignment = Alignment(horizontal="center")

    pretty = {
        "price_per_unit": "Price per unit",
        "variable_cost_per_unit": "Variable cost / unit",
        "fixed_cost": "Fixed cost",
    }
    row = 2
    for var in SENSITIVITY_VARIABLES:
        for r in [s for s in result.sensitivity if s.variable == var]:
            ws.cell(row=row, column=1, value=pretty.get(var, var))
            ws.cell(row=row, column=2, value=r.delta_pct).number_format = PCT_FMT
            ws.cell(row=row, column=3, value=r.new_value).number_format = fmt
            if r.new_break_even_units is None:
                ws.cell(row=row, column=4, value="n/a")
                ws.cell(row=row, column=5, value="n/a")
                ws.cell(row=row, column=6, value="n/a")
            else:
                ws.cell(row=row, column=4, value=r.new_break_even_units).number_format = INT_FMT
                if r.swing_units is not None:
                    ws.cell(row=row, column=5, value=r.swing_units).number_format = INT_FMT
                if r.swing_pct is not None:
                    ws.cell(row=row, column=6, value=r.swing_pct).number_format = PCT_FMT
            if r.delta_pct == 0.0:
                for col in range(1, 7):
                    ws.cell(row=row, column=col).fill = PANEL_FILL
            row += 1

    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 18

    if result.tornado:
        png_bytes = render_tornado_png(result.tornado, title="Tornado: BE swing per variable (+/-10%)")
        img = XLImage(BytesIO(png_bytes))
        img.width, img.height = 600, 280
        ws.add_image(img, "H2")
